"""
Document Ingestion Layer
Handles multiple document types: Native PDF, Scanned PDF, Excel
Normalizes all outputs to unified format for Groq extraction
Uses EasyOCR (no system dependencies needed)
"""

import PyPDF2
import io
import os
from typing import List, Dict
from pathlib import Path

# OCR imports (will fail gracefully if not installed)
try:
    import fitz  # PyMuPDF
    import easyocr
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("⚠️ OCR not available. Install: pip install -r requirements.txt")

# Excel imports
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Global EasyOCR reader (loaded once)
_ocr_reader = None

def _get_ocr_reader():
    """Load EasyOCR reader once (expensive operation)"""
    global _ocr_reader
    if _ocr_reader is None and OCR_AVAILABLE:
        try:
            print("📦 Loading EasyOCR model (one-time initialization)...")
            _ocr_reader = easyocr.Reader(['en'], gpu=False)
            print("✅ EasyOCR model loaded")
        except Exception as e:
            print(f"⚠️ Failed to load EasyOCR: {e}")
            return None
    return _ocr_reader


def detect_file_type(filename: str) -> str:
    """Detect file type from extension"""
    ext = Path(filename).suffix.lower()
    
    if ext == ".pdf":
        return "PDF"
    elif ext in [".xlsx", ".xls"]:
        return "Excel"
    elif ext in [".docx", ".doc"]:
        return "Word"
    else:
        return "Unknown"


def is_scanned_pdf(pdf_bytes: bytes) -> bool:
    """
    Detect if PDF is scanned (image-only) or native (text selectable)
    
    Returns:
        True if scanned/image-only PDF
        False if native PDF with selectable text
    """
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
        
        # Check first 5 pages for substantial text
        for page_num in range(min(5, len(pdf_reader.pages))):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            
            # If we find substantial text, it's a native PDF
            if text and len(text.strip()) > 100:
                print(f"✅ Page {page_num + 1}: Found {len(text.strip())} characters of text → NATIVE PDF")
                return False
        
        # No substantial text found → it's scanned
        print("ℹ️ No substantial text found in PDF pages → SCANNED PDF detected")
        return True
    except Exception as e:
        print(f"⚠️ Error checking PDF: {e}")
        # Default to OCR if error occurs
        return True


def extract_with_pypdf(pdf_bytes: bytes) -> List[Dict]:
    """
    Extract text from native PDF using PyPDF2
    
    Returns:
        List of dicts with keys: page_number, text, source_type, character_count
    """
    pages_text = []
    
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
        
        for page_num, page in enumerate(pdf_reader.pages, start=1):
            text = page.extract_text()
            pages_text.append({
                'page_number': page_num,
                'text': text if text else '',
                'source_type': 'PDF_Native',
                'character_count': len(text) if text else 0
            })
        
        print(f"✅ PyPDF2 extracted {len(pages_text)} pages")
        return pages_text
    except Exception as e:
        print(f"❌ Error extracting PDF with PyPDF2: {e}")
        return []


def extract_with_ocr(pdf_bytes: bytes, dpi: int = 300) -> List[Dict]:
    """
    Extract text from scanned PDF using EasyOCR
    No system dependencies needed - works on Windows, Linux, macOS
    
    Args:
        pdf_bytes: PDF file bytes
        dpi: Resolution for rendering (higher = better quality, slower)
    
    Returns:
        List of dicts with keys: page_number, text, source_type, character_count
    """
    if not OCR_AVAILABLE:
        print("❌ OCR not available. Install: pip install easyocr PyMuPDF")
        return []
    
    pages_text = []
    
    try:
        # Initialize EasyOCR reader
        reader = _get_ocr_reader()
        if not reader:
            print("❌ Failed to initialize EasyOCR reader")
            return []
        
        # Open PDF with PyMuPDF (no poppler needed)
        print(f"🖼️ Converting PDF to images using PyMuPDF...")
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_pages = len(pdf_document)
        print(f"📖 Processing {total_pages} pages with EasyOCR...")
        
        for page_num in range(total_pages):
            try:
                # Render page to image at specified DPI
                page = pdf_document[page_num]
                
                # Scale factor for DPI (default 72 DPI = 1.0, 300 DPI = ~4.17)
                zoom = dpi / 72.0
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                
                # Convert to numpy array for EasyOCR
                import numpy as np
                img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                
                # Run OCR
                print(f"  Page {page_num + 1}: Running EasyOCR...")
                results = reader.readtext(img_array, detail=0)
                text = '\n'.join(results) if results else ''
                
                pages_text.append({
                    'page_number': page_num + 1,
                    'text': text,
                    'source_type': 'PDF_Scanned',
                    'character_count': len(text),
                    'dpi': dpi
                })
                
                if text:
                    print(f"  ✅ Page {page_num + 1}: {len(text)} characters extracted")
                else:
                    print(f"  ⚠️ Page {page_num + 1}: No text detected")
                
            except Exception as e:
                print(f"  ⚠️ Page {page_num + 1} error: {e}")
                pages_text.append({
                    'page_number': page_num + 1,
                    'text': '',
                    'source_type': 'PDF_Scanned',
                    'character_count': 0,
                    'dpi': dpi
                })
                continue
        
        pdf_document.close()
        print(f"✅ EasyOCR extracted {len(pages_text)} pages")
        return pages_text
        
    except Exception as e:
        print(f"❌ Error extracting with OCR: {e}")
        import traceback
        traceback.print_exc()
        return []


def extract_with_excel(file_bytes: bytes) -> List[Dict]:
    """
    Extract requirements from Excel spreadsheet
    
    Args:
        file_bytes: Excel file bytes (.xlsx or .xls)
    
    Returns:
        List of dicts with keys: page_number, text, source_type, sheet_name, row_number, character_count
    """
    if not EXCEL_AVAILABLE:
        print("❌ Excel support not available. Install: pip install openpyxl")
        return []
    
    items = []
    page_counter = 0
    
    try:
        print(f"📊 Opening Excel workbook...")
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
        sheet_names = workbook.sheetnames
        print(f"✅ Found {len(sheet_names)} sheet(s): {sheet_names}")
        
        # Process each sheet
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            print(f"\n📋 Processing sheet: '{sheet_name}'")
            
            row_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                # Skip completely empty rows
                if all(cell.value is None for cell in row):
                    continue
                
                row_count += 1
                page_counter += 1
                
                # Extract cell values from this row
                cell_values = []
                for cell in row:
                    value = cell.value
                    
                    if value is None:
                        cell_values.append('')
                    elif isinstance(value, bool):
                        cell_values.append('Yes' if value else 'No')
                    elif isinstance(value, (int, float)):
                        cell_values.append(str(value))
                    else:
                        cell_values.append(str(value).strip())
                
                # Create formatted text for this row
                row_text = ' | '.join(cell_values)
                
                items.append({
                    'page_number': page_counter,
                    'text': row_text,
                    'source_type': 'Excel',
                    'sheet_name': sheet_name,
                    'row_number': row_idx,
                    'character_count': len(row_text)
                })
                
                if row_count % 5 == 0 or row_count == 1:
                    preview = row_text[:60] + ('...' if len(row_text) > 60 else '')
                    print(f"  Row {row_idx}: {preview}")
            
            print(f"  ✅ Extracted {row_count} non-empty rows")
        
        workbook.close()
        print(f"\n✅ Excel extraction complete: {len(items)} total rows")
        return items
        
    except Exception as e:
        print(f"❌ Error extracting Excel: {e}")
        import traceback
        traceback.print_exc()
        return []


def normalize_document(pages_or_items: List[Dict]) -> str:
    """
    Convert any format to normalized text with metadata
    
    Args:
        pages_or_items: List of dicts with 'text' and page/sheet metadata
    
    Returns:
        Single string with [PAGE n] or [SHEET name] markers preserved
    """
    normalized_parts = []
    
    for item in pages_or_items:
        source = item.get('source_type', 'Unknown')
        
        if source == 'Excel':
            header = f"\n[EXCEL: {item.get('sheet_name', 'Unknown')} - Row {item.get('row_number', 1)}]\n"
        elif source == 'Word':
            header = f"\n[WORD: {item.get('section_name', 'Document')}]\n"
        else:
            # PDF sources
            header = f"\n[{source} - Page {item.get('page_number', 1)}]\n"
        
        text = item.get('text', '')
        normalized_parts.append(header + text)
    
    return "".join(normalized_parts)


def chunk_document(normalized_text: str, chunk_size: int = 2000) -> List[Dict]:
    """
    Break normalized document into chunks while preserving page references
    
    Args:
        normalized_text: Output from normalize_document()
        chunk_size: Max characters per chunk
    
    Returns:
        List of chunks with page range tracking
    """
    chunks = []
    current_chunk = ""
    current_page = 1
    
    for line in normalized_text.split('\n'):
        # Track page changes
        if line.startswith('['):
            # Extract page number if present
            if 'Page' in line:
                try:
                    page_num = int(line.split('Page')[-1].split(']')[0].strip())
                    current_page = page_num
                except:
                    pass
        
        # Add line to current chunk
        current_chunk += line + '\n'
        
        # Split if chunk is too large
        if len(current_chunk) > chunk_size:
            chunks.append({
                'text': current_chunk,
                'page_range': f"{current_page}",
                'character_count': len(current_chunk)
            })
            current_chunk = ""
    
    # Add remaining
    if current_chunk:
        chunks.append({
            'text': current_chunk,
            'page_range': f"{current_page}",
            'character_count': len(current_chunk)
        })
    
    return chunks


# Public API
__all__ = [
    'detect_file_type',
    'is_scanned_pdf',
    'extract_with_pypdf',
    'extract_with_ocr',
    'extract_with_excel',
    'normalize_document',
    'chunk_document'
]