"""
Excel Export Module for PFMEA Reports
Generates formatted Excel files with FMEA data, RPN scores, and summary
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.database import fetch_one, fetch_all
from app.rpn_suggestion_engine import get_rpn_summary


def export_pfmea_to_excel(part_id: int, file_path: str) -> str:
    """
    Generate Excel file for PFMEA record with all failure modes and scores
    
    Args:
        part_id: ID of the PFMEA record to export
        file_path: Path where to save the Excel file
    
    Returns:
        Path to saved file
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "PFMEA"
    
    # ========== Fetch Data ==========
    part = fetch_one("SELECT * FROM pfmea_records WHERE id = %s", (part_id,))
    if not part:
        raise ValueError(f"PFMEA record with ID {part_id} not found")
    
    entries = fetch_all("""
        SELECT 
            pfe.*,
            fm.canonical_name as failure_mode_name,
            fm.category as failure_mode_category,
            ps.step_name as process_step_name
        FROM pfmea_failure_mode_entries pfe
        JOIN failure_mode_taxonomy fm ON pfe.failure_mode_id = fm.id
        LEFT JOIN process_steps ps ON pfe.process_step_id = ps.id
        WHERE pfe.pfmea_record_id = %s
        ORDER BY pfe.process_step_number
    """, (part_id,))
    
    # Fetch causes and controls for each entry
    for entry in entries:
        entry['causes'] = fetch_all("""
            SELECT * FROM failure_mode_causes WHERE fmea_entry_id = %s
            ORDER BY cause_sequence
        """, (entry['id'],))
        
        entry['controls'] = fetch_all("""
            SELECT * FROM process_controls WHERE fmea_entry_id = %s
            ORDER BY control_type DESC
        """, (entry['id'],))
    
    # ========== Set Column Widths ==========
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 5
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    
    # ========== Header Section ==========
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    cells_to_merge = [
        ('A', 'J', row)
    ]
    
    ws.merge_cells(f'A{row}:J{row}')
    cell = ws[f'A{row}']
    cell.value = part.get('part_name', 'PFMEA Report')
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    # Metadata rows
    metadata = [
        ("Part Number", part.get('part_number')),
        ("Model Year", part.get('model_year')),
        ("Part Name", part.get('part_name')),
        ("Customer", part.get('customer_name')),
        ("Process Responsibility", part.get('process_responsibility')),
        ("FMEA Date (Original)", part.get('fmea_date_original')),
        ("Format Number", part.get('format_number')),
        ("Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = str(value) if value else ""
        row += 1
    
    row += 1  # Blank row
    
    # ========== Main FMEA Table ==========
    headers = [
        "Step #",
        "Process Step",
        "Failure Mode",
        "Potential Effect",
        "Severity",
        "Occurrence",
        "Detection",
        "Risk Priority Number",
        "Suggested RPN",
        "Risk Level"
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = border
    
    row += 1
    
    # ========== Data Rows ==========
    data_fill_odd = PatternFill(start_color="E7F0F7", end_color="E7F0F7", fill_type="solid")
    data_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for idx, entry in enumerate(entries):
        fill = data_fill_even if idx % 2 == 0 else data_fill_odd
        
        # Get scores
        severity = entry.get('severity_user_input') or entry.get('severity_suggested') or ""
        occurrence = entry.get('occurrence_user_input') or entry.get('occurrence_suggested') or ""
        detection = entry.get('detection_user_input') or entry.get('detection_suggested') or ""
        rpn = entry.get('rpn_user_calculated') or entry.get('rpn_suggested') or ""
        suggested = entry.get('rpn_suggested') or ""
        
        # Determine risk class and color
        if rpn:
            if rpn > 70:
                risk_class = "HIGH"
                risk_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                risk_font = Font(bold=True, color="FFFFFF", size=11)
            elif rpn >= 40:
                risk_class = "MED"
                risk_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                risk_font = Font(bold=True, color="000000", size=11)
            else:
                risk_class = "LOW"
                risk_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                risk_font = Font(bold=True, color="FFFFFF", size=11)
        else:
            risk_class = ""
            risk_fill = fill
            risk_font = Font()
        
        # Row data
        row_data = [
            entry.get('process_step_number'),
            entry.get('process_step_name'),
            entry.get('failure_mode_name'),
            entry.get('potential_effect'),
            severity,
            occurrence,
            detection,
            rpn,
            suggested,
            risk_class
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Apply fill
            if col_idx < 10:
                cell.fill = fill
            else:
                cell.fill = risk_fill
            
            # Apply font to risk column
            if col_idx == 10:
                cell.font = risk_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx >= 5:  # S, O, D, RPN, Suggested columns
                cell.alignment = Alignment(horizontal="center", vertical="top")
        
        # Add notes row if canvas_notes exist
        if entry.get('canvas_notes'):
            row += 1
            cell = ws[f'A{row}']
            cell.value = f"Note: {entry.get('canvas_notes')}"
            cell.font = Font(italic=True, size=9)
            ws.merge_cells(f'A{row}:J{row}')
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        row += 1
    
    # ========== Summary Sheet ==========
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = get_rpn_summary(entries)
    
    row = 1
    ws_summary[f'A{row}'] = "FMEA Component Summary"
    ws_summary[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 2
    
    # Summary metrics
    summary_metrics = [
        ("Part Number", part.get('part_number')),
        ("Part Name", part.get('part_name')),
        ("", ""),
        ("Max RPN (Highest Risk)", summary_data['max']),
        ("Average RPN", f"{summary_data['average']:.1f}"),
        ("Total Failure Modes", summary_data['total_failure_modes']),
        ("", ""),
        ("Risk Classification", "Count"),
        ("HIGH (RPN > 70)", summary_data['high_count']),
        ("MED (RPN 40-70)", summary_data['med_count']),
        ("LOW (RPN < 40)", summary_data['low_count']),
    ]
    
    for label, value in summary_metrics:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True) if label != "" else Font()
        
        if label in ["Max RPN (Highest Risk)", "HIGH (RPN > 70)", "MED (RPN 40-70)", "LOW (RPN < 40)"]:
            if label == "Max RPN (Highest Risk)":
                ws_summary[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws_summary[f'B{row}'].font = Font(bold=True, color="FFFFFF", size=11)
            elif label == "HIGH (RPN > 70)":
                ws_summary[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws_summary[f'B{row}'].font = Font(bold=True, color="FFFFFF")
            elif label == "MED (RPN 40-70)":
                ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                ws_summary[f'B{row}'].font = Font(bold=True)
            elif label == "LOW (RPN < 40)":
                ws_summary[f'B{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                ws_summary[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        
        ws_summary[f'B{row}'] = value if value != "" else ""
        row += 1
    
    # Set column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    
    # Save workbook
    wb.save(file_path)
    return file_path
