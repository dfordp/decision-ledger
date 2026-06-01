"""
Generate ReviewGraph Platform Capability Matrix Excel
10 columns including Evaluation Input Stream and Verdict/Output Type
based on the three-stream AI evaluation model:
  Rule Engine (amber) | Part Requirements (blue) | Historical Intelligence (purple)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Core palette ─────────────────────────────────────────────────────────────
C_HEADER_BG   = "1B2A4A"
C_HEADER_FG   = "FFFFFF"
C_LAYER_BG    = [
    "1D3557",   # 1 – Drawing Ingestion
    "1A3A4A",   # 2 – Eng. Graph
    "2D4739",   # 3 – Revision Comparison
    "3B2D52",   # 4 – Deterministic Val.
    "3B3018",   # 5 – AI Reasoning
    "1F3B5C",   # 6 – Reviewer Workflow
    "1A2E1A",   # 7 – Final Outputs
]
C_LAYER_FG    = "FFFFFF"
C_ROW_EVEN    = ["EBF1FB","E6F0F5","EBF5EE","F0EBF8","FDF5E6","E8EFF7","EBF5EB"]
C_ROW_ODD     = ["FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF"]

C_AI_ALONE    = "D1FAE5";  C_CHECK_AI    = "065F46"
C_AI_JOINT    = "DBEAFE";  C_CHECK_JOINT = "1E3A8A"
C_CLIENT      = "FEF3C7";  C_CHECK_CLI   = "92400E"

THIN = Side(style="thin",   color="D0D7E3")
MED  = Side(style="medium", color="9BA8B7")

# ── Evaluation stream palette (from image prompt) ─────────────────────────────
# Rule Engine      = amber/gold    (the deterministic rule cascade)
# Part Requirements = electric blue (ECN history, revision lineage, specs)
# Historical Intel. = deep violet   (reviewer patterns, prior decisions)
# Human Decision    = charcoal      (sign-off, override, judgment)
# Composite         = soft neutral  (multiple streams converging)

STREAM_STYLE = {
    "RE":    ("FEF3C7", "92400E", "Rule Engine"),
    "PR":    ("DBEAFE", "1E3A8A", "Part Requirements"),
    "HI":    ("EDE9FE", "5B21B6", "Historical Intelligence"),
    "HD":    ("F3F4F6", "374151", "Human Decision"),
    "RE+PR": ("FFF7CD", "78350F", "Rule Engine · Part Requirements"),
    "HI+PR": ("E5DEFF", "4C1D95", "Historical Intelligence · Part Requirements"),
    "RE+HI": ("FEF9C3", "713F12", "Rule Engine · Historical Intelligence"),
    "ALL":   ("F0F4FF", "1E3A5F", "All Three Streams"),
}

# Verdict output style
VERDICT_STYLE = {
    "BLOCK/WARN/SAFE": ("FEF2F2", "991B1B"),   # red tint
    "WARN/SAFE":       ("FFFBEB", "92400E"),   # amber tint
    "Dashboard":       ("EFF6FF", "1E40AF"),   # blue tint
    "Graph":           ("F0FDF4", "166534"),   # green tint
    "Narrative":       ("F8FAFC", "475569"),   # slate
    "Structured Data": ("F0F9FF", "0C4A6E"),   # sky blue
    "Impact Map":      ("FFF7ED", "9A3412"),   # orange
    "Stage Flag":      ("ECFDF5", "065F46"),   # emerald
    "Export":          ("F5F3FF", "5B21B6"),   # purple
    "Human Input":     ("F9FAFB", "6B7280"),   # gray
    "Override Record": ("FFFBEB", "78350F"),   # warm amber
    "Weighted Score":  ("F0FDF4", "14532D"),   # deep green
    "Advisory":        ("EFF6FF", "1D4ED8"),   # indigo
}

def vstyle(key):
    return VERDICT_STYLE.get(key, ("FFFFFF","374151"))

def sstyle(key):
    return STREAM_STYLE.get(key, ("FFFFFF","374151","Unknown"))

# ── Data ─────────────────────────────────────────────────────────────────────
# Tuple: (layer, capability, ai_alone, ai+client, client_only,
#         reasoning, example, prod_notes, stream_key, verdict_key)
ROWS = [

    # ══ Layer 1 · Drawing Ingestion Pipeline ══════════════════════════════════
    (1,
     "Multi-format ingestion & format normalisation\n(PNG / PDF / DXF / DWG / CAD)",
     True, False, False,
     "File parsing rules are fully deterministic — format conversion and structural "
     "normalisation require no domain knowledge.",
     "Upload HB-000235-R01.png → system normalises raster image to structured "
     "extraction data via _EXTRACTION_TEMPLATES lookup.",
     "Supports PNG, PDF, DXF, DWG; OCR degradation on hand-annotated or <300 DPI scans.",
     "PR", "Structured Data"),

    (1,
     "OCR scanning & title-block text extraction",
     True, False, False,
     "Standard engineering drawing layouts follow fixed zones — pattern matching is reliable.",
     "HB-000235 title block: Part No, Rev, Scale 1:1, Material CRCA IS 513 Grade D extracted automatically.",
     "Accuracy degrades on non-standard title blocks; client must supply zone-mapping config for bespoke formats.",
     "RE", "Structured Data"),

    (1,
     "GD&T symbol parsing\n(feature control frames, datum references)",
     True, False, False,
     "ISO / ASME Y14.5 GD&T notation is machine-readable with rule-based parsers.",
     "Ø0.25 A, Ø0.10 M B callouts extracted from HB-000235; H11 fit class parsed from hole callout strings.",
     "Requires model calibration for non-standard or legacy annotation styles.",
     "RE", "Structured Data"),

    (1,
     "BOM / ERP / PLM metadata linking",
     False, True, False,
     "AI identifies drawing fields but client must confirm correct BOM hierarchy and ERP line mapping.",
     "HB-000235 part number matched to Hatchback C vehicle program; client confirms Escorts Kubota PLM mapping.",
     "Requires client ERP connector; field mapping varies by ERP (SAP, Teamcenter, etc.).",
     "PR", "Structured Data"),

    (1,
     "Geometry parsing — 2-D topology, section views",
     True, False, False,
     "Shape recognition from 2-D projections is ML-driven and operates without client input.",
     "Mounting hole topology extracted from HB-000235 section view; 5 holes mapped to upper/centre/lower regions.",
     "Parse quality varies with drawing clarity and projection completeness.",
     "PR", "Structured Data"),

    # ══ Layer 2 · Structured Engineering Graph ════════════════════════════════
    (2,
     "Feature extraction\n(holes, slots, chamfers, weld notes)",
     True, False, False,
     "Geometric and annotation features are pattern-matched from extracted drawing data without domain input.",
     "HB-000235-R01: 5 holes (2xO13 H11, 1xO18.1 H11, 2xO12 H11), slot 26x12 mm, E-coat finish extracted.",
     "Feature count forms basis for MOUNTING_HOLE_COUNT_REQUIRED rule; mismatch resolved via artifact rule profile.",
     "RE+PR", "Graph"),

    (2,
     "Dimension graph construction\n(spatial relationships between nodes)",
     True, False, False,
     "Numerical values and units extracted deterministically; graph edges computed from spatial co-references.",
     "Overall height 152.1 mm, hole centre distance 51.4 mm, side reference height 89.9 mm mapped as graph nodes.",
     "Tolerance bands defined in post-Groq reconciliation _required_dims_all_present().",
     "RE", "Graph"),

    (2,
     "Tolerance & fit-class capture",
     True, False, False,
     "Standard tolerance notation (H11, ISO 2768-m) parsed deterministically from hole callout strings.",
     "_has_h11 computed by scanning all hole callouts for 'H11' — drives ENGINE_LOAD_FIT_CLASS verdict.",
     "Post-Groq reconciliation overrides any hallucinated BLOCK findings when _has_h11 is True.",
     "RE", "BLOCK/WARN/SAFE"),

    (2,
     "Assembly interface & mates-and-fits mapping",
     False, True, False,
     "AI suggests interfaces from BOM links; client engineer confirms mating part and assembly hierarchy.",
     "Bracket Engine Load (HB-000235) interfaces with engine mount sub-assembly — client confirms BOM parent.",
     "Requires confirmed BOM; incorrect parent causes erroneous impact-propagation results in Layer 3.",
     "PR", "Graph"),

    (2,
     "Revision lineage graph construction",
     True, False, False,
     "Sequential revision uploads automatically build the supersession chain — no human input needed.",
     "R00 → R01 lineage for HB-000235 built automatically from upload order; displayed in Revision Comparison Engine.",
     "Lineage integrity depends on uploads being performed in correct revision order.",
     "HI", "Graph"),

    (2,
     "BOM link validation & consistency check",
     False, True, False,
     "AI cross-references drawing metadata with PLM data; client validates correct part-family assignment.",
     "HB-000235 (Hatchback C) vs HB-000071 (Hatchback A) — client confirms these are distinct vehicle programs.",
     "Critical for search stage-flag accuracy (has_approved_revision via BOOL_OR SQL).",
     "PR", "Graph"),

    # ══ Layer 3 · Revision Comparison Engine ══════════════════════════════════
    (3,
     "Delta extraction\n(dimension / tolerance / note changes R-to-R)",
     True, False, False,
     "Computed from structured graph node diffing — no domain knowledge required beyond the graph data.",
     "R00 to R01: O18.1 H11 centre hole added (ECN-235-003); process changed to Laser Cut + Bend; fatigue note added.",
     "Accuracy depends on Layer 2 graph completeness; missing nodes produce false-negative deltas.",
     "HI", "Narrative"),

    (3,
     "Change summary generation",
     True, False, False,
     "Template-driven narrative from extracted deltas — no human input needed for standard change types.",
     "Auto-summary: 'Prototype release for Hatchback C — all critical dimensions confirmed, O18.1 H11 per ECN-235-003 added.'",
     "Stored in change_summary field of _EXTRACTION_TEMPLATES; overwrites on re-upload.",
     "HI", "Narrative"),

    (3,
     "Impact propagation\n(assembly, manufacturing, BOM, NVH)",
     False, True, False,
     "AI computes blast radius from graph relationships; client engineer confirms impact priority and downstream action owners.",
     "O18.1 hole addition triggers NVH-HBC-001 validation plan reference — client confirms whether NVH sign-off is a release gate.",
     "Impact scope depends on completeness of BOM links; partial BOM = under-reported impact.",
     "RE+PR", "Impact Map"),

    (3,
     "Cost & lead-time impact estimation",
     False, False, True,
     "Requires client supplier pricing models, capacity data, and lead-time agreements — entirely outside AI knowledge.",
     "Process change Stamping to Laser Cut + Bend: client supplies tooling cost delta and revised lead time.",
     "Not currently implemented in ReviewGraph v1; placeholder field in impact propagation output.",
     "HD", "Human Input"),

    (3,
     "Delta visualisation\n(added / removed / changed overlay)",
     True, False, False,
     "Colour-coded overlay is generated automatically from the delta set — pure rendering logic.",
     "Green = O18.1 hole added; amber = process annotation changed; rendered in Revision Comparison table.",
     "Visual output rendered in reviewer workspace revision comparison table.",
     "HI", "Dashboard"),

    # ══ Layer 4 · Deterministic Validation Engine ══════════════════════════════
    (4,
     "Drafting-standard validation\n(line standards, annotation, GD&T, title block)",
     True, False, False,
     "Rules encoded per ASME Y14.5, ISO 2768, DIN EN — fully deterministic with explainable pass/fail per rule.",
     "ISO 2768-m general tolerance stated on HB-000235-R01 → GENERAL_TOLERANCE_STANDARD_STATED: SAFE.",
     "500+ rules in engine; HB-000235 profile skips this as SKIP to avoid false WARN on non-standard layout.",
     "RE", "BLOCK/WARN/SAFE"),

    (4,
     "Manufacturing feasibility\n(machinability, DFM, process capability)",
     True, False, False,
     "Rule-based against known process capabilities encoded in validation engine; no real-time supplier data needed.",
     "Laser Cut + Bend valid for CRCA 2.5 mm sheet stock — process feasibility SAFE for HB-000235-R01.",
     "Rules are static; new process types require rule additions by engineering team.",
     "RE", "BLOCK/WARN/SAFE"),

    (4,
     "Tolerance stack-up & fit-class validation\n(H11, m6, ISO fits)",
     True, False, False,
     "Deterministic: _has_h11=True → ENGINE_LOAD_FIT_CLASS SAFE; _has_h11=False → BLOCK. "
     "Post-Groq reconciliation enforces this regardless of LLM output.",
     "HB-000235-R01 all 5 holes carry H11 → _has_h11=True → all three H11 rule keys forced SAFE in reconciliation pass.",
     "Reconciliation dict _H11_RULES: ENGINE_LOAD_FIT_CLASS, UPPER_HOLE_DIAMETER_PROTECTED, LOWER_HOLE_DIAMETER_PROTECTED.",
     "RE", "BLOCK/WARN/SAFE"),

    (4,
     "Fatigue / load-spectrum note validation",
     True, False, False,
     "Pre-computed: any of ('PT-LS', 'fatigue', 'load spectrum', 'NVH') in drawing notes → FATIGUE_SPECTRUM_NOTE_REQUIRED: SAFE.",
     "HB-000235-R01 notes contain 'FATIGUE LOAD SPECTRUM: PT-LS-235' → _has_fatigue_note=True → SAFE, no BLOCK.",
     "HB-000235-R00 notes = 'First issue — not yet released' (no fatigue keywords) → BLOCK correctly demonstrated.",
     "RE", "BLOCK/WARN/SAFE"),

    (4,
     "Artifact rule profile overrides\n(SKIP / WARN / BLOCK per artifact)",
     False, False, True,
     "Client domain experts define per-artifact exceptions that override universal rule set — requires deep part knowledge.",
     "HB-000235: MOUNTING_HOLE_COUNT_REQUIRED → SKIP (part has 5 holes, rule expects 4). 33 SKIPs, 8 WARNs, 7 BLOCKs defined.",
     "Seeded via _seed_hb235_rule_profile() in seed_escorts_demo.py; persisted in design_artifact_rules table.",
     "HD", "BLOCK/WARN/SAFE"),

    (4,
     "Inspection readiness\n(CTQ identification, gage plan, sampling rules)",
     False, True, False,
     "AI flags missing inspection references; client provides actual inspection plan and CTQ definitions.",
     "FATIGUE_SPECTRUM_NOTE_REQUIRED cross-checks against PT-LS-235 note; client confirms gage plan covers O18.1 H11 hole.",
     "Gage availability and sampling rules require client quality engineering input.",
     "RE+PR", "WARN/SAFE"),

    (4,
     "BOM consistency & part-family validation",
     False, True, False,
     "AI cross-references rule scope against BOM hierarchy; client confirms correct part family grouping for rule inheritance.",
     "HB-000235 belongs to Bracket Engine Load family — client confirms rules from HB-000071 family apply as baseline.",
     "Rule inheritance across part families reduces re-configuration effort on similar parts.",
     "RE+PR", "WARN/SAFE"),

    # ══ Layer 5 · AI Reasoning Layer ══════════════════════════════════════════
    (5,
     "Reviewer pattern learning\n(behaviour, preferences, decision patterns)",
     True, False, False,
     "Learns from historical approval and override records — entirely data-driven with no manual configuration.",
     "System learns that ECN-numbered hole callouts on Bracket Engine Load family are always H11-validated — reduces future false BLOCKs.",
     "Requires sufficient override history (10+ decisions) to generalise; cold-start risk on new artifact families.",
     "HI", "Weighted Score"),

    (5,
     "Engineering context understanding\n(intent, function, manufacturing context)",
     False, True, False,
     "LLM infers design intent from notes and callouts; client validates that inferred context matches actual engineering intent.",
     "NVH-HBC-001 plan reference recognised as fatigue-critical context — client confirms this is a release gate.",
     "Groq llama-4-scout-17b-16e-instruct used; hallucination risk mitigated by post-Groq Python reconciliation.",
     "HI+PR", "Advisory"),

    (5,
     "Semantic drawing understanding\n(LLM parsing of free-text notes and callouts)",
     True, False, False,
     "LLM parses engineering notes semantically without client input — pre-computed facts constrain the output.",
     "'WIRE ROUTING SLOT 26x12 mm PER ECN-235-001. DEBURR ALL SLOT EDGES.' → slot function and edge treatment understood.",
     "Groq prompt injects pre-computed _fit_class_verdict and _notes_verdict to prevent model overriding deterministic facts.",
     "RE+PR", "Advisory"),

    (5,
     "Contextual validation memory\n(storing prior decisions for similar parts)",
     True, False, False,
     "Prior validation decisions stored in drawing_validation_results and referenced for similar artifact families.",
     "Previous H11 approval pattern on HB-000071 Hatchback A informs confidence scoring for HB-000235 Hatchback C.",
     "Memory is per-artifact-family; cross-family generalisation requires explicit similarity mapping.",
     "HI", "Weighted Score"),

    (5,
     "Dynamic review recommendations\n(highlight risks, suggest checks)",
     False, True, False,
     "AI surfaces recommendations; reviewer decides which to act on — prioritisation requires domain judgment.",
     "'Verify NVH-HBC-001 plan sign-off before approving R01 — fatigue note present but plan not yet approved.'",
     "Recommendations rendered in reviewer workspace; dismissal is logged for pattern learning.",
     "HI", "Advisory"),

    (5,
     "Engineering intent mapping\n(design intent to validation outcome)",
     False, True, False,
     "AI proposes function-to-feature mapping; client confirms design intent before encoded as a validation rule.",
     "Centre hole O18.1 H11 mapped to wire routing function per ECN-235-003 — client confirms, system encodes SAFE rationale.",
     "Intent mapping drives which rule keys apply; incorrect intent = wrong SKIP/BLOCK assignment.",
     "HI+PR", "Advisory"),

    # ══ Layer 6 · Reviewer Workflow ════════════════════════════════════════════
    (6,
     "Reviewer dashboard population\n(findings summary, revision status)",
     True, False, False,
     "Automated aggregation of validation run results — no human input needed to populate the dashboard.",
     "HB-000235-R01 post-reconciliation: 0 BLOCK, 3 WARN, 14+ SAFE findings displayed automatically after upload.",
     "Dashboard data sourced from drawing_validation_results table; reflects latest run only unless history view enabled.",
     "RE+HI", "Dashboard"),

    (6,
     "Explainable findings with rule evidence",
     True, False, False,
     "Each finding carries rule_key, status, evidence snippet, and standard reference — generated deterministically.",
     "FATIGUE_SPECTRUM_NOTE_REQUIRED: status=SAFE, evidence='PT-LS-235 found in drawing notes', standard='ASME Y14.5'.",
     "Explainability is a core platform guarantee: same input → same rules → same output, always traceable.",
     "RE", "BLOCK/WARN/SAFE"),

    (6,
     "Evidence graph generation\n(node path from feature to verdict)",
     True, False, False,
     "Graph traversal from feature node through rule node to verdict node is automated from Layer 2 graph data.",
     "O18.1 H11 hole node → ENGINE_LOAD_FIT_CLASS rule node → SAFE verdict path rendered in evidence graph panel.",
     "Graph completeness depends on Layer 2 extraction quality.",
     "RE+HI", "Graph"),

    (6,
     "Reviewer notes & annotations",
     False, False, True,
     "Free-text engineering judgment cannot be automated — requires reviewer domain expertise and accountability.",
     "'Approved with condition: NVH-HBC-001 plan sign-off required before SOP.' — entered by lead engineer A. Patel.",
     "Notes stored against revision; surfaced in audit trail and downstream PLM export.",
     "HD", "Narrative"),

    (6,
     "Override & decision recording\n(WARN to SAFE, BLOCK to SAFE with justification)",
     False, True, False,
     "AI presents the finding and evidence; client engineer provides justification and accepts accountability.",
     "Lead engineer overrides SURFACE_FINISH_REQUIRED WARN → SAFE: 'E-coat per HB-SPEC-021 confirmed by supplier audit.'",
     "Full audit trail: who overrode, when, justification, linked evidence. Feeds reviewer pattern learning in Layer 5.",
     "HD", "Override Record"),

    (6,
     "Signoff lifecycle management\n(In Review to Approved to Released)",
     False, True, False,
     "AI tracks state transitions and enforces gate conditions; human engineers perform each sign-off action.",
     "HB-000235-R01: In Review → Engineering Review (0 BLOCK confirmed) → Approved → has_approved_revision=True.",
     "approval_status field drives has_approved_revision via BOOL_OR SQL aggregate in search_design_artifacts().",
     "HD", "Stage Flag"),

    (6,
     "Production / In-Development stage flag\n(part search visual indicator)",
     True, False, False,
     "Fully computed: BOOL_OR(approval_status IN ('approved','approved_reference')) → has_approved_revision → green/amber dot.",
     "HB-000071 & HB-000110: Mass Production (green dot). HB-000235: In Development (amber dot) until R01 approved.",
     "Bug fixed: /api/part-search dict must explicitly include has_approved_revision — omission caused all cards to show In Development.",
     "HI", "Stage Flag"),

    # ══ Layer 7 · Final Outputs ════════════════════════════════════════════════
    (7,
     "Release readiness dashboard\n(% ready, at-risk, not-ready breakdown)",
     True, False, False,
     "Computed from open BLOCK and WARN counts against total rule set — no human input needed.",
     "72% readiness: 0 BLOCK open, 3 WARN pending reviewer judgment, 14 SAFE passed for HB-000235-R01.",
     "Readiness threshold for release gate is client-configurable (default: 0 BLOCK, 5 or fewer WARN).",
     "RE+HI", "Dashboard"),

    (7,
     "Engineering validation report (PDF export)",
     True, False, False,
     "Auto-generated from validation run data — rule verdicts, evidence, override log, approval chain.",
     "Full report for HB-000235-R01: all 17 rule verdicts, 3 override records, lead engineer sign-off, NVH condition note.",
     "PDF generation via WeasyPrint / ReportLab; template configurable per client quality system.",
     "RE", "Narrative"),

    (7,
     "Manufacturing risk summary\n(risk level, impact by category)",
     True, False, False,
     "Aggregated from fabrication, assembly, and tolerance validation results — risk level computed deterministically.",
     "HB-000235-R01: MEDIUM risk — Laser Cut + Bend process change, NVH plan not yet closed.",
     "Risk matrix thresholds (Low/Medium/High) are client-configurable.",
     "RE+PR", "Impact Map"),

    (7,
     "Revision impact analysis\n(components affected, BOM impact, lead time)",
     True, False, False,
     "Delta + impact propagation combined into structured impact report — AI-generated from Layer 3 outputs.",
     "R00 to R01: 3 components affected (bracket, wire harness clip, NVH mount), NVH validation plan triggered.",
     "Lead-time and cost columns remain client-filled; see Layer 3 note.",
     "HI+PR", "Impact Map"),

    (7,
     "Approval decision\n(Conditional Approval / Full Approval / Rejected)",
     False, False, True,
     "Final approval is always a human decision — AI provides all evidence but cannot sign off on behalf of an engineer.",
     "'CONDITIONAL APPROVAL — R01 Hatchback C prototype: NVH-HBC-001 plan sign-off required before SOP.' — R. Menon.",
     "Approval recorded in drawing_revisions.approval_status; drives has_approved_revision and stage flag.",
     "HD", "Human Input"),

    (7,
     "PLM / ERP / QMS / MES export\n(Teamcenter, SAP, ETQ, Siemens)",
     False, True, False,
     "AI formats the export payload; client configures connector credentials, field mapping, and approval triggers.",
     "On R01 conditional approval: export to Teamcenter PLM (part revision record) and ETQ QMS (CAPA reference).",
     "Connector configuration is client-specific; ReviewGraph provides API webhooks and structured JSON output.",
     "PR", "Export"),
]

# ── Layer metadata ────────────────────────────────────────────────────────────
LAYER_NAMES = {
    1: "1 · Drawing Ingestion Pipeline",
    2: "2 · Structured Engineering Graph",
    3: "3 · Revision Comparison Engine",
    4: "4 · Deterministic Validation Engine",
    5: "5 · AI Reasoning Layer",
    6: "6 · Reviewer Workflow",
    7: "7 · Final Outputs",
}
LAYER_SUB = {
    1: "Multi-source Ingestion & Intelligent Parsing",
    2: "Engineering Knowledge Graph & Relationships",
    3: "Multi-Revision Delta Analysis & Impact Propagation",
    4: "Rule-Based · Deterministic · Explainable",
    5: "Contextual Intelligence · Learning · Reasoning",
    6: "Human-in-the-Loop Engineering Review",
    7: "Insights · Reports · Decisions",
}

# ── Column layout  A–J ────────────────────────────────────────────────────────
HEADERS = [
    "Platform Layer",                       # A
    "Capability / Workflow",                # B
    "AI Can Handle\nIndependently",         # C
    "Requires AI +\nClient / Reviewer",     # D
    "Requires Client /\nDomain Knowledge",  # E
    "Reasoning",                            # F
    "Example From Workflow",                # G
    "Production Notes",                     # H
    "Evaluation Input\nStream",             # I  ← NEW
    "Verdict /\nOutput Type",               # J  ← NEW
]
COL_WIDTHS = [22, 36, 10, 11, 11, 44, 50, 44, 22, 17]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Capability Matrix"

for i, w in enumerate(COL_WIDTHS):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# ── Row 1 – title ─────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 36
ws.merge_cells("A1:J1")
t = ws["A1"]
t.value = "ReviewGraph  —  Platform Capability Matrix"
t.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
t.fill = PatternFill("solid", fgColor="0F1929")
t.alignment = Alignment(horizontal="center", vertical="center")

# ── Row 2 – subtitle ──────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 18
ws.merge_cells("A2:J2")
s = ws["A2"]
s.value = ("AI-Assisted Engineering Drawing Review & Revision Intelligence Platform   |   "
           "Three-Stream Evaluation Model:  Rule Engine  ·  Part Requirements  ·  Historical Intelligence")
s.font = Font(name="Calibri", italic=True, size=9.5, color="A0AEC0")
s.fill = PatternFill("solid", fgColor="1B2A4A")
s.alignment = Alignment(horizontal="center", vertical="center")

# ── Row 3 – stream legend bar ─────────────────────────────────────────────────
ws.row_dimensions[3].height = 20

label_cfg = [
    (1, 3, "Rule Engine", "FEF3C7", "92400E",
     "Amber stream — 500+ deterministic rules, ASME Y14.5, ISO 2768, fit-class gates"),
    (4, 6, "Part Requirements", "DBEAFE", "1E3A8A",
     "Blue stream — ECN history, revision lineage, part specs, BOM"),
    (7, 9, "Historical Intelligence", "EDE9FE", "5B21B6",
     "Violet stream — reviewer patterns, prior approvals, institutional memory"),
    (10, 10, "Human Decision", "F3F4F6", "374151",
     "Charcoal — sign-off, override, domain judgment"),
]

for (c1, c2, label, bg, fg, tip) in label_cfg:
    start = get_column_letter(c1)
    end   = get_column_letter(c2)
    if c1 != c2:
        ws.merge_cells(f"{start}3:{end}3")
    cell = ws.cell(row=3, column=c1)
    cell.value = label
    cell.font  = Font(name="Calibri", bold=True, size=9, color=fg)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Row 4 – column headers ────────────────────────────────────────────────────
ws.row_dimensions[4].height = 36
for col_idx, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=4, column=col_idx, value=hdr)
    cell.font = Font(name="Calibri", bold=True, size=9.5, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1B2A4A")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="medium", color="4A90D9"),
        right =Side(style="thin",   color="2D4A7A"),
    )

# ── Data rows ─────────────────────────────────────────────────────────────────
current_layer = None
row_num = 5

def cell_border():
    return Border(
        left  =Side(style="thin", color="D0D7E3"),
        right =Side(style="thin", color="D0D7E3"),
        top   =Side(style="thin", color="E5E9F0"),
        bottom=Side(style="thin", color="D0D7E3"),
    )

for (layer, cap, ai, joint, cli, reason, example, notes, stream_key, verdict_key) in ROWS:
    bg_idx = layer - 1

    # ── Layer banner row ──────────────────────────────────────────────────────
    if layer != current_layer:
        ws.row_dimensions[row_num].height = 26
        ws.merge_cells(f"A{row_num}:J{row_num}")
        lh = ws.cell(row=row_num, column=1)
        lh.value = f"  {LAYER_NAMES[layer]}   —   {LAYER_SUB[layer]}"
        lh.font = Font(name="Calibri", bold=True, size=10.5, color=C_LAYER_FG)
        lh.fill = PatternFill("solid", fgColor=C_LAYER_BG[bg_idx])
        lh.alignment = Alignment(horizontal="left", vertical="center")
        lh.border = Border(
            top   =Side(style="medium", color="FFFFFF"),
            bottom=Side(style="medium", color="FFFFFF"),
        )
        row_num += 1
        current_layer = layer

    # ── Data row ──────────────────────────────────────────────────────────────
    is_even = (row_num % 2 == 0)
    row_bg  = C_ROW_EVEN[bg_idx] if is_even else C_ROW_ODD[bg_idx]
    ws.row_dimensions[row_num].height = 58

    # A – blank (layer shown in banner)
    c = ws.cell(row=row_num, column=1, value="")
    c.fill = PatternFill("solid", fgColor=row_bg)
    c.border = cell_border()

    # B – Capability
    c = ws.cell(row=row_num, column=2, value=cap)
    c.font = Font(name="Calibri", bold=True, size=9.5, color="0F1929")
    c.fill = PatternFill("solid", fgColor=row_bg)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border = cell_border()

    # C – AI alone
    c = ws.cell(row=row_num, column=3, value="✓" if ai else "")
    c.font = Font(name="Calibri", bold=True, size=13,
                  color=C_CHECK_AI if ai else "D0D7E3")
    c.fill = PatternFill("solid", fgColor=C_AI_ALONE if ai else row_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border()

    # D – AI + client
    c = ws.cell(row=row_num, column=4, value="✓" if joint else "")
    c.font = Font(name="Calibri", bold=True, size=13,
                  color=C_CHECK_JOINT if joint else "D0D7E3")
    c.fill = PatternFill("solid", fgColor=C_AI_JOINT if joint else row_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border()

    # E – Client only
    c = ws.cell(row=row_num, column=5, value="✓" if cli else "")
    c.font = Font(name="Calibri", bold=True, size=13,
                  color=C_CHECK_CLI if cli else "D0D7E3")
    c.fill = PatternFill("solid", fgColor=C_CLIENT if cli else row_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border()

    # F – Reasoning
    c = ws.cell(row=row_num, column=6, value=reason)
    c.font = Font(name="Calibri", size=9, color="374151")
    c.fill = PatternFill("solid", fgColor=row_bg)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border = cell_border()

    # G – Example
    c = ws.cell(row=row_num, column=7, value=example)
    c.font = Font(name="Calibri", size=9, color="1E3A8A")
    c.fill = PatternFill("solid", fgColor=row_bg)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border = cell_border()

    # H – Production Notes
    c = ws.cell(row=row_num, column=8, value=notes)
    c.font = Font(name="Calibri", size=9, color="6B7A90", italic=True)
    c.fill = PatternFill("solid", fgColor=row_bg)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border = cell_border()

    # I – Evaluation Input Stream  ← NEW
    sbg, sfg, slabel = sstyle(stream_key)
    c = ws.cell(row=row_num, column=9, value=slabel)
    c.font = Font(name="Calibri", bold=True, size=8.5, color=sfg)
    c.fill = PatternFill("solid", fgColor=sbg)
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    c.border = Border(
        left  =Side(style="medium", color=sfg),
        right =Side(style="thin",   color="D0D7E3"),
        top   =Side(style="thin",   color="E5E9F0"),
        bottom=Side(style="thin",   color="D0D7E3"),
    )

    # J – Verdict / Output  ← NEW
    vbg, vfg = vstyle(verdict_key)
    c = ws.cell(row=row_num, column=10, value=verdict_key)
    c.font = Font(name="Calibri", bold=True, size=9, color=vfg)
    c.fill = PatternFill("solid", fgColor=vbg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(
        left  =Side(style="thin",   color="D0D7E3"),
        right =Side(style="medium", color=vfg),
        top   =Side(style="thin",   color="E5E9F0"),
        bottom=Side(style="thin",   color="D0D7E3"),
    )

    row_num += 1

# ── Legend ─────────────────────────────────────────────────────────────────────
row_num += 1
ws.row_dimensions[row_num].height = 38

# Section title
ws.merge_cells(f"A{row_num}:B{row_num}")
c = ws.cell(row=row_num, column=1, value="Legend — Ownership")
c.font = Font(name="Calibri", bold=True, size=9, color="374151")
c.alignment = Alignment(vertical="center")

for col_idx, (val, bg, fg) in enumerate([
    ("AI Independently", C_AI_ALONE,  C_CHECK_AI),
    ("AI + Client",      C_AI_JOINT,  C_CHECK_JOINT),
    ("Client / Domain",  C_CLIENT,    C_CHECK_CLI),
], start=3):
    c = ws.cell(row=row_num, column=col_idx, value=f"✓  {val}")
    c.font  = Font(name="Calibri", size=9, bold=True, color=fg)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Stream legend
row_num += 1
ws.row_dimensions[row_num].height = 38
ws.merge_cells(f"A{row_num}:B{row_num}")
c = ws.cell(row=row_num, column=1, value="Legend — Evaluation Stream")
c.font = Font(name="Calibri", bold=True, size=9, color="374151")
c.alignment = Alignment(vertical="center")

stream_legend_items = [
    ("Rule Engine",              "FEF3C7", "92400E"),
    ("Part Requirements",        "DBEAFE", "1E3A8A"),
    ("Historical Intelligence",  "EDE9FE", "5B21B6"),
    ("Human Decision",           "F3F4F6", "374151"),
]
for i, (label, bg, fg) in enumerate(stream_legend_items):
    col = 3 + i
    c = ws.cell(row=row_num, column=col, value=label)
    c.font  = Font(name="Calibri", size=9, bold=True, color=fg)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Notes merged across F:J
ws.merge_cells(f"F{row_num - 1}:J{row_num}")
note_c = ws.cell(row=row_num - 1, column=6,
    value=("Three-stream evaluation model from ReviewGraph AI prompt:  "
           "Rule Engine (amber) — 500+ deterministic standards · ASME Y14.5 · ISO 2768 · fit-class gates.  "
           "Part Requirements (blue) — ECN history · revision lineage · engineering change notices · part specs.  "
           "Historical Intelligence (violet) — reviewer patterns · prior approvals · institutional memory · "
           "Groq LLM contextual reasoning.  "
           "All examples: HB-000071 (Hatchback A, Mass Production), HB-000110 (Hatchback B, Mass Production), "
           "HB-000235 (Hatchback C, In Development). "
           "Deterministic Guarantee: Same Input -> Same Rules -> Same Output, always."))
note_c.font = Font(name="Calibri", size=8, italic=True, color="8492A6")
note_c.alignment = Alignment(wrap_text=True, vertical="top")

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = "A5"

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"D:\Projects\Development\decision-ledger\ReviewGraph_Capability_Matrix.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total rows written up to: {row_num}")
